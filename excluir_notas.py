"""
Limpa duplicatas do controle.xlsx, mantendo para cada (data, tipo_nota)
apenas a linha com o melhor status.

Ordem de prioridade de status (maior = melhor):
  CONCLUIDO > NOTA_CRIADA > ANEXANDO > PENDENTE > ERRO
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook

DATA_DIR = Path(os.environ.get("DATA_DIR", "/data"))
CONTROLE_PATH = DATA_DIR / "controle" / "controle.xlsx"

COLUNAS = [
    "data", "dia", "tipo_nota", "numero_nota",
    "arquivo_escala", "status", "observacao", "processado_em",
]

PRIORIDADE = {
    "CONCLUIDO":   5,
    "NOTA_CRIADA": 4,
    "ANEXANDO":    3,
    "PENDENTE":    2,
    "ERRO":        1,
}


def _prioridade(status: str) -> int:
    return PRIORIDADE.get((status or "").strip().upper(), 0)


def limpar():
    if not CONTROLE_PATH.exists():
        print(f"Arquivo não encontrado: {CONTROLE_PATH}")
        sys.exit(1)

    wb = load_workbook(CONTROLE_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("Arquivo vazio.")
        sys.exit(0)

    header = [str(c).strip() if c else "" for c in rows[0]]

    linhas = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        valores = [str(v).strip() if v is not None else "" for v in row]
        linhas.append(dict(zip(header, valores)))

    total_antes = len(linhas)

    # Deduplicar: para cada (data, tipo_nota) mantém a linha com maior prioridade
    mapa: dict = {}
    for linha in linhas:
        chave = (linha.get("data", ""), linha.get("tipo_nota", ""))
        existente = mapa.get(chave)
        if existente is None or _prioridade(linha.get("status")) > _prioridade(existente.get("status")):
            mapa[chave] = linha

    unicas = list(mapa.values())

    # Reordena por data e tipo
    def sort_key(r):
        d, mo, a = (r.get("data") or "01/01/2000").split("/")
        return (int(a), int(mo), int(d), r.get("tipo_nota", ""))

    unicas.sort(key=sort_key)

    total_depois = len(unicas)
    removidas = total_antes - total_depois

    # Salva
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "controle"
    ws2.append(COLUNAS)
    for linha in unicas:
        ws2.append([linha.get(col, "") for col in COLUNAS])
    wb2.save(CONTROLE_PATH)

    print(f"Antes : {total_antes} linhas")
    print(f"Depois: {total_depois} linhas")
    print(f"Removidas: {removidas} duplicatas")
    print()
    for linha in unicas:
        print(f"  {linha.get('data'):12} | {linha.get('tipo_nota'):15} | {linha.get('status'):12} | nota: {linha.get('numero_nota')}")


if __name__ == "__main__":
    limpar()
