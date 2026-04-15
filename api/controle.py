"""
Leitura e escrita do arquivo de controle (controle.xlsx).
É a fonte da verdade do lote: o que processar, status atual, número da nota.

Colunas:
  data | dia | tipo_nota | numero_nota | arquivo_escala | status | observacao | processado_em

Status possíveis:
  PENDENTE → NOTA_CRIADA → ANEXANDO → CONCLUIDO
                                     → ERRO
"""
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook, load_workbook

COLUNAS = [
    "data",
    "dia",
    "tipo_nota",
    "numero_nota",
    "arquivo_escala",
    "status",
    "observacao",
    "processado_em",
]

STATUS_FINAL = {"CONCLUIDO"}
STATUS_REPROCESSAR = {"ERRO", "PENDENTE"}


def _controle_path() -> Path:
    base = Path(os.environ.get("DATA_DIR", "/data"))
    return base / "controle" / "controle.xlsx"


def _criar_arquivo_vazio(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "controle"
    ws.append(COLUNAS)
    wb.save(path)


def ler_controle() -> List[Dict]:
    """Lê todas as linhas do controle. Retorna lista de dicts."""
    path = _controle_path()
    if not path.exists():
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    resultado = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        valores = [str(v).strip() if v is not None else "" for v in row]
        item = dict(zip(header, valores))
        resultado.append(item)

    return resultado


def salvar_controle(linhas: List[Dict]):
    """Sobrescreve o arquivo de controle com as linhas fornecidas."""
    path = _controle_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "controle"
    ws.append(COLUNAS)

    for linha in linhas:
        row = [linha.get(col, "") for col in COLUNAS]
        ws.append(row)

    wb.save(path)


def adicionar_linhas(novas: List[Dict]) -> int:
    """
    Adiciona linhas ao controle apenas se não existir linha com
    mesma (data, tipo_nota) em status que não seja ERRO.
    Retorna quantas linhas foram de fato adicionadas.
    """
    existentes = ler_controle()
    adicionadas = 0

    for nova in novas:
        chave_data = nova.get("data", "")
        chave_tipo = nova.get("tipo_nota", "")

        # Verifica idempotência: já existe qualquer linha para esta data+tipo?
        ja_existe = any(
            l.get("data") == chave_data
            and l.get("tipo_nota") == chave_tipo
            for l in existentes
        )

        if not ja_existe:
            nova.setdefault("status", "PENDENTE")
            nova.setdefault("numero_nota", "")
            nova.setdefault("observacao", "")
            nova.setdefault("processado_em", "")
            existentes.append(nova)
            adicionadas += 1

    salvar_controle(existentes)
    return adicionadas


def atualizar_linha(data: str, tipo_nota: str, **campos):
    """
    Atualiza campos de uma linha existente identificada por (data, tipo_nota).
    Campos aceitos: status, numero_nota, observacao, processado_em.
    """
    linhas = ler_controle()
    for linha in linhas:
        if linha.get("data") == data and linha.get("tipo_nota") == tipo_nota:
            for k, v in campos.items():
                if k in COLUNAS:
                    linha[k] = v
            break
    salvar_controle(linhas)


def marcar_status(data: str, tipo_nota: str, status: str, **extra):
    """Atalho para atualizar status e campos opcionais."""
    campos = {"status": status, "processado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    campos.update(extra)
    atualizar_linha(data, tipo_nota, **campos)


def buscar(data: Optional[str] = None, status_list: Optional[List[str]] = None) -> List[Dict]:
    """
    Retorna linhas filtradas por data e/ou status.
    Se nenhum filtro for passado, retorna tudo.
    """
    linhas = ler_controle()

    if data:
        linhas = [l for l in linhas if l.get("data") == data]

    if status_list:
        linhas = [l for l in linhas if l.get("status") in status_list]

    return linhas
