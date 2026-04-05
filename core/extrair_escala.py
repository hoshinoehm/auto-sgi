"""
Wrapper de extração de escala com output_dir configurável.
Toda a lógica de leitura/escrita é a mesma de escalas/março/extrair_escala_para_excel.py.
A diferença é que aqui o diretório de saída é um parâmetro explícito.

Função pública principal:
  extrair_arquivo(input_path, output_dir) -> Dict
"""
import re
import zipfile
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constantes (idênticas ao script original)
# ---------------------------------------------------------------------------

COLS = ["NOME", "MATRÍCULA", "ID", "FUNÇÃO", "OPÇÔES", "TURNO / QTU"]
ADMIN_TURNO = "EXPEDIENTE ADMINISTRITIVO"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=11, bold=False)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

OPCOES_POR_FUNCAO = {
    "CPU": "CPU",
    "CMT GU": "Cmt / Guarnição",
    "PAT": "Patrulheiro",
    "MOT": "Motorista Vtr",
    "CMT OM": "Administrativo (Cmt)",
    "SCMT": "Administrativo (S Cmt)",
    "S CMT": "Administrativo (S Cmt)",
    "P1": "Administrativo (P/1)",
    "P/1": "Administrativo (P/1)",
    "P2": "Administrativo (P/2)",
    "P/2": "Administrativo (P/2)",
    "P3": "Administrativo (P/3)",
    "P/3": "Administrativo (P/3)",
    "P4": "Administrativo (P/4)",
    "P/4": "Administrativo (P/4)",
    "AUX P1": "Aux Administrativo (P/1)",
    "AUX P/1": "Aux Administrativo (P/1)",
    "AUX P2": "Aux Administrativo (P/2)",
    "AUX P/2": "Aux Administrativo (P/2)",
    "AUX P3": "Aux Administrativo (P/3)",
    "AUX P/3": "Aux Administrativo (P/3)",
    "AUX P4": "Aux Administrativo (P/4)",
    "AUX P/4": "Aux Administrativo (P/4)",
    "PERM": "Permanência",
}

# ---------------------------------------------------------------------------
# Helpers (idênticos ao script original)
# ---------------------------------------------------------------------------

def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", ("" if s is None else str(s))).strip()


def _upper(s: Any) -> str:
    return _norm(s).upper()


def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")


def _norm_key(s: Any) -> str:
    return re.sub(r"\s+", " ", _strip_accents(_upper(s)))


def _should_ignore(nome: str) -> bool:
    return re.search(r"\bALONSO\b", _norm_key(nome)) is not None


def _fix_matricula_id(matricula: str, pid: str) -> Tuple[str, str]:
    m = _norm(matricula).replace("…", "...").strip()
    i = _norm(pid).replace("…", "...").strip()
    if _upper(m) in {"", "..."}:
        m = i
    if _upper(i) in {"", "..."}:
        i = m
    return m, i


def _opcoes_por_funcao(funcao: str) -> str:
    return OPCOES_POR_FUNCAO.get(_norm_key(funcao), "")


def _is_operacional(funcao: str) -> bool:
    return _norm_key(funcao) in {"CPU", "CMT GU", "PAT", "MOT"}


# ---------------------------------------------------------------------------
# Leitura de DOCX (idêntica ao script original)
# ---------------------------------------------------------------------------

def _read_docx(path: Path) -> Dict[str, List[Dict]]:
    from lxml import etree

    zf = zipfile.ZipFile(path)
    xml_bytes = zf.read("word/document.xml")
    parser = etree.XMLParser(huge_tree=True, recover=True)
    root = etree.fromstring(xml_bytes, parser=parser)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    def extract_tables():
        tables = []
        for tbl in root.xpath(".//w:tbl", namespaces=ns):
            t = []
            for tr in tbl.xpath("./w:tr", namespaces=ns):
                row = []
                for tc in tr.xpath("./w:tc", namespaces=ns):
                    cell = "".join([t.text for t in tc.xpath(".//w:t", namespaces=ns) if t.text])
                    cell = re.sub(r"\s+", " ", cell).strip()
                    row.append(cell)
                if any(_norm(c) != "" for c in row):
                    t.append(row)
            if t:
                tables.append(t)
        return tables

    tables = extract_tables()
    all_rows: List[Dict] = []

    for tbl in tables:
        header_idx = None
        for i, row in enumerate(tbl[:10]):
            joined = "|".join(_norm_key(x) for x in row)
            if "NOME" in joined and "ID" in joined:
                header_idx = i
                break
        if header_idx is None:
            continue

        header = [_norm_key(x) for x in tbl[header_idx]]

        def idx(*names: str) -> Optional[int]:
            names_u = [_norm_key(n) for n in names]
            for j, h in enumerate(header):
                if h in names_u:
                    return j
            for j, h in enumerate(header):
                if any(n in h for n in names_u):
                    return j
            return None

        i_nome = idx("NOME")
        i_mat = idx("MATRÍCULA", "MATRICULA")
        i_id = idx("ID")
        i_fun = idx("FUNÇÃO", "FUNCAO")
        i_turno = idx("TURNO / QTU", "TURNO", "QTU")

        section = tbl[0][0] if tbl and tbl[0] else ""

        for row in tbl[header_idx + 1:]:
            nome = row[i_nome] if (i_nome is not None and i_nome < len(row)) else ""
            if _norm(nome) == "":
                continue
            if _should_ignore(nome):
                continue

            matricula = row[i_mat] if (i_mat is not None and i_mat < len(row)) else ""
            pid = row[i_id] if (i_id is not None and i_id < len(row)) else ""
            funcao = row[i_fun] if (i_fun is not None and i_fun < len(row)) else ""

            if _norm(funcao) == "" and "CPU" in _norm_key(section):
                funcao = "CPU"

            matricula, pid = _fix_matricula_id(matricula, pid)

            turno = ""
            if i_turno is not None and i_turno < len(row):
                turno = _norm(row[i_turno])

            all_rows.append({
                "NOME": _upper(nome),
                "MATRÍCULA": matricula,
                "ID": pid,
                "FUNÇÃO": _upper(funcao),
                "OPÇÔES": _opcoes_por_funcao(funcao),
                "TURNO / QTU": turno,
            })

    administrativo = []
    operacional = []
    for r in all_rows:
        if _is_operacional(r["FUNÇÃO"]):
            if _norm(r["TURNO / QTU"]) == "":
                r["TURNO / QTU"] = "24h"
            operacional.append(r)
        else:
            r["TURNO / QTU"] = ADMIN_TURNO
            administrativo.append(r)

    return {"ADMINISTRATIVO": administrativo, "OPERACIONAL": operacional}


# ---------------------------------------------------------------------------
# Leitura de XLSX (idêntica ao script original)
# ---------------------------------------------------------------------------

def _read_xlsx(path: Path) -> Dict[str, List[Dict]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    def find_cell_with(text: str):
        target = _norm_key(text)
        for row in ws.iter_rows():
            for cell in row:
                if _norm_key(cell.value) == target:
                    return cell.row, cell.column
        return None

    def read_block(start_row, start_col, is_admin):
        header_row = start_row + 1
        headers = []
        col = start_col
        while True:
            v = _norm(ws.cell(row=header_row, column=col).value)
            if v == "":
                break
            headers.append(v)
            col += 1

        hmap = {_norm_key(h): i for i, h in enumerate(headers)}

        def get(row_vals, *keys):
            for k in keys:
                ku = _norm_key(k)
                if ku in hmap and hmap[ku] < len(row_vals):
                    return _norm(row_vals[hmap[ku]])
            return ""

        out = []
        r = header_row + 1
        while True:
            row_vals = [ws.cell(row=r, column=start_col + i).value for i in range(len(headers))]
            row_vals = [_norm(v) for v in row_vals]
            nome = get(row_vals, "NOME")
            if nome == "":
                break
            if _should_ignore(nome):
                r += 1
                continue

            funcao = get(row_vals, "FUNÇÃO", "FUNCAO")
            matricula = get(row_vals, "MATRÍCULA", "MATRICULA")
            pid = get(row_vals, "ID")
            turno = get(row_vals, "TURNO / QTU", "TURNO/ QTU", "TURNO")
            matricula, pid = _fix_matricula_id(matricula, pid)

            if is_admin:
                turno = ADMIN_TURNO
            else:
                if _is_operacional(funcao) and _norm(turno) == "":
                    turno = "24h"

            out.append({
                "NOME": _upper(nome),
                "MATRÍCULA": matricula,
                "ID": pid,
                "FUNÇÃO": _upper(funcao),
                "OPÇÔES": _opcoes_por_funcao(funcao),
                "TURNO / QTU": turno,
            })
            r += 1
        return out

    result = {"ADMINISTRATIVO": [], "OPERACIONAL": []}
    pos_adm = find_cell_with("ADMINISTRATIVO")
    if pos_adm:
        result["ADMINISTRATIVO"] = read_block(*pos_adm, True)
    pos_op = find_cell_with("OPERACIONAL")
    if pos_op:
        result["OPERACIONAL"] = read_block(*pos_op, False)
    return result


# ---------------------------------------------------------------------------
# Sanitização (idêntica ao script original)
# ---------------------------------------------------------------------------

def _sanitize(data: Dict) -> Dict:
    out = {"ADMINISTRATIVO": [], "OPERACIONAL": []}
    for grupo in ["ADMINISTRATIVO", "OPERACIONAL"]:
        for r in data.get(grupo, []):
            if _should_ignore(r.get("NOME", "")):
                continue
            r["NOME"] = _upper(r.get("NOME", ""))
            r["FUNÇÃO"] = _upper(r.get("FUNÇÃO", ""))
            r["OPÇÔES"] = _opcoes_por_funcao(r.get("FUNÇÃO", ""))
            m, i = _fix_matricula_id(r.get("MATRÍCULA", ""), r.get("ID", ""))
            r["MATRÍCULA"] = m
            r["ID"] = i
            if grupo == "ADMINISTRATIVO":
                r["TURNO / QTU"] = ADMIN_TURNO
            else:
                if _is_operacional(r.get("FUNÇÃO", "")) and _norm(r.get("TURNO / QTU", "")) == "":
                    r["TURNO / QTU"] = "24h"
            out[grupo].append(r)
    return out


# ---------------------------------------------------------------------------
# Escrita do XLSX de saída (idêntica ao script original)
# ---------------------------------------------------------------------------

def _autosize(ws):
    for col_idx, col_name in enumerate(COLS, start=1):
        max_len = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)


def _apply_style(ws, title: str):
    ws.title = "Plan1"
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].border = THIN_BORDER
    for col_idx, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER


def _write_xlsx(rows: List[Dict], title: str, output_path: Path):
    wb = Workbook()
    ws = wb.active
    _apply_style(ws, title)
    for i, r in enumerate(rows, start=3):
        for col_idx, col_name in enumerate(COLS, start=1):
            ws.cell(row=i, column=col_idx, value=r.get(col_name, "")).font = DATA_FONT
    _autosize(ws)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Nomes dos arquivos de saída
# ---------------------------------------------------------------------------

def _output_names(input_path: Path, output_dir: Path) -> Tuple[Path, Path]:
    """
    Retorna (operacional_path, administrativo_path) em output_dir.
    Padrão: ESCALA - 19.02.2026 - QUINTA-FEIRA.docx →
      operacional:     ESCALA - 19.02.2026 - QUINTA-FEIRA_EXTRAIDA.xlsx
      administrativo:  ESCALA - 19.02.2026 - QUINTA-FEIRA_ADM_EXTRAIDA.xlsx
    """
    stem = input_path.stem
    stem = re.sub(r"_EXTRAIDA$", "", stem, flags=re.IGNORECASE)

    base = re.sub(
        r"^LANÇAR ESCALA\s*-\s*ADM\s*-\s*",
        "LANÇAR ESCALA - ",
        stem,
        flags=re.IGNORECASE,
    )

    if re.match(r"^LANÇAR ESCALA\s*-\s*", base, flags=re.IGNORECASE):
        op_name = f"{base}_EXTRAIDA.xlsx"
        adm_name = re.sub(
            r"^LANÇAR ESCALA\s*-\s*", "LANÇAR ESCALA - ADM - ", base, flags=re.IGNORECASE
        ) + "_EXTRAIDA.xlsx"
    else:
        op_name = f"{base}_EXTRAIDA.xlsx"
        adm_name = f"{base}_ADM_EXTRAIDA.xlsx"

    return output_dir / op_name, output_dir / adm_name


# ---------------------------------------------------------------------------
# Função pública
# ---------------------------------------------------------------------------

def extrair_arquivo(input_path: Path, output_dir: Path) -> Dict:
    """
    Extrai uma escala de DOCX (ou XLSX) e salva dois arquivos em output_dir:
      - *_EXTRAIDA.xlsx         (operacional)
      - *_ADM_EXTRAIDA.xlsx     (administrativo)

    Retorna:
      {
        "operacional": {"caminho": str, "total": int},
        "administrativo": {"caminho": str, "total": int},
      }

    Lança exceção em caso de falha na leitura.
    """
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    suffix = input_path.suffix.lower()
    if suffix == ".docx":
        data = _read_docx(input_path)
    elif suffix == ".xlsx":
        data = _read_xlsx(input_path)
    else:
        raise ValueError(f"Formato não suportado: {suffix}. Use .docx ou .xlsx")

    data = _sanitize(data)

    op_path, adm_path = _output_names(input_path, output_dir)

    _write_xlsx(data.get("OPERACIONAL", []), "OPERACIONAL", op_path)
    _write_xlsx(data.get("ADMINISTRATIVO", []), "ADMINISTRATIVO", adm_path)

    print(f"[EXTRAÇÃO] Entrada : {input_path.name}")
    print(f"[EXTRAÇÃO] Operacional    ({len(data['OPERACIONAL'])} registros): {op_path.name}")
    print(f"[EXTRAÇÃO] Administrativo ({len(data['ADMINISTRATIVO'])} registros): {adm_path.name}")

    return {
        "operacional": {
            "caminho": str(op_path),
            "arquivo": op_path.name,
            "total_registros": len(data["OPERACIONAL"]),
        },
        "administrativo": {
            "caminho": str(adm_path),
            "arquivo": adm_path.name,
            "total_registros": len(data["ADMINISTRATIVO"]),
        },
    }
