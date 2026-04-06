"""
Lógica de anexação de registros de escala a uma nota no SGI.
Extraída de sgi_anexos_por_nota.py — sem input(), sem getpass().
Credenciais vêm por parâmetro (vindos de variáveis de ambiente no contexto da API).

Função pública principal:
  anexar_lote(nota_id, xlsx_path, usuario, senha) -> Dict
"""
import os
import re
import time
import subprocess
import unicodedata
from pathlib import Path
from typing import List, Dict

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoAlertPresentException
from openpyxl import load_workbook

from core.driver import criar_driver
from core.sgi_auth import login_completo

URL_ANEXOS_BASE = "https://sgi.pm.ma.gov.br/boletim_eletronico_dp_anexos_incluir.php?ID="
DEFAULT_TIMEOUT = int(os.environ.get("SELENIUM_TIMEOUT", "30"))


# ---------------------------------------------------------------------------
# Normalização de texto (mesma lógica de sgi_anexos_por_nota.py)
# ---------------------------------------------------------------------------

def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")


def _norm_txt(s: str) -> str:
    s = (s or "").strip()
    s = _strip_accents(s)
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def _norm_header(s: str) -> str:
    s = (s or "").strip()
    s = _strip_accents(s)
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def _canon_turno(turno: str) -> str:
    t = (turno or "").strip()
    if not t:
        return t
    t = re.sub(r"ADMINISTRITIVO", "ADMINISTRATIVO", t, flags=re.IGNORECASE)
    if _norm_txt(t) in {"24h", "24 h"}:
        return "24h"
    return t


# ---------------------------------------------------------------------------
# Leitura da planilha (mesma lógica de sgi_anexos_por_nota.py)
# ---------------------------------------------------------------------------

def _ler_registros_xlsx(caminho: str) -> List[Dict]:
    print(f"[XLSX] Lendo: {caminho}")
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    header_row_idx = None
    headers_norm = None

    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if not row or all(v is None for v in row):
            continue

        row_raw = [str(v).strip() if v is not None else "" for v in row]
        row_norm = [_norm_header(v) for v in row_raw]
        hset = set(row_norm)

        has_matricula = "MATRICULA" in hset or "MATRÍCULA" in row_raw
        has_opcoes = any(x in hset for x in {"OPCOES", "OPÇÕES", "OPÇÔES"})
        has_turno = any(x in hset for x in {"TURNO / QTU", "TURNO/QTU", "TURNO", "QTU"})

        if has_matricula and has_opcoes and has_turno:
            header_row_idx = idx
            headers_norm = row_norm
            print(f"[XLSX] Cabeçalho encontrado na linha {idx}")
            break

    if header_row_idx is None:
        print("[XLSX] Cabeçalho não encontrado (MATRÍCULA, OPÇÕES, TURNO).")
        return []

    def _col(*nomes):
        candidatos = {_norm_header(n) for n in nomes}
        for i, h in enumerate(headers_norm):
            if h in candidatos:
                return i
        return None

    idx_mat = _col("MATRÍCULA", "MATRICULA")
    idx_tipo = _col("OPÇÕES", "OPÇÔES", "OPCOES")
    idx_turno = _col("TURNO / QTU", "TURNO/QTU", "TURNO", "QTU")

    if None in (idx_mat, idx_tipo, idx_turno):
        print("[XLSX] Não encontrei todos os índices das colunas necessárias.")
        return []

    dados = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        if not row or all(v is None for v in row):
            continue

        valores = [str(v).strip() if v is not None else "" for v in row]
        if all(v == "" for v in valores):
            continue

        if max(idx_mat, idx_tipo, idx_turno) >= len(valores):
            continue

        matricula = valores[idx_mat]
        tipo_servico = valores[idx_tipo]
        turno = _canon_turno(valores[idx_turno])

        if not matricula or not tipo_servico or not turno:
            continue

        dados.append({"matricula": matricula, "tipo_servico": tipo_servico, "turno": turno})

    print(f"[XLSX] {len(dados)} registros carregados.")
    return dados


# ---------------------------------------------------------------------------
# Selenium: navegação e preenchimento (mesma lógica de sgi_anexos_por_nota.py)
# ---------------------------------------------------------------------------

def _selecionar_option_relaxado(select: Select, texto: str, label: str = "") -> bool:
    desejado = _norm_txt(texto)
    for opt in select.options:
        if opt.text and _norm_txt(opt.text) == desejado:
            select.select_by_visible_text(opt.text)
            return True
    for opt in select.options:
        if opt.text and desejado and desejado in _norm_txt(opt.text):
            select.select_by_visible_text(opt.text)
            return True
    for opt in select.options:
        if opt.text and desejado and _norm_txt(opt.text).startswith(desejado):
            select.select_by_visible_text(opt.text)
            return True
    opcoes = [opt.text for opt in select.options if opt.text]
    print(f"[DEBUG] Opção '{texto}' não encontrada em '{label}'. Disponíveis: {opcoes}")
    return False


def _tratar_alerta_sucesso(driver):
    try:
        WebDriverWait(driver, 5).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        print(f"[ALERTA] {alert.text}")
        alert.accept()
        time.sleep(1)
    except (TimeoutException, NoAlertPresentException):
        pass


def _ir_para_anexos(driver, nota_id: str) -> bool:
    wait = WebDriverWait(driver, DEFAULT_TIMEOUT)
    url = URL_ANEXOS_BASE + nota_id.strip()
    driver.get(url)
    try:
        wait.until(lambda d: "boletim_eletronico_dp_anexos_incluir.php" in d.current_url)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "form")))
        return True
    except TimeoutException:
        print(f"[ERRO] Não carregou página de anexos para nota {nota_id}")
        return False


def _preencher_registro(driver, dados: Dict, indice: int, total: int):
    wait = WebDriverWait(driver, DEFAULT_TIMEOUT)

    matricula = dados["matricula"]
    tipo_servico = dados["tipo_servico"]
    turno = _canon_turno(dados["turno"])

    print(f"[{indice}/{total}] MAT={matricula} | TIPO={tipo_servico} | TURNO={turno}")

    mat_input = wait.until(EC.presence_of_element_located((By.NAME, "matricula")))
    mat_input.clear()
    mat_input.send_keys(matricula)

    finalidade_el = wait.until(EC.presence_of_element_located((By.NAME, "finalidade")))
    if not _selecionar_option_relaxado(Select(finalidade_el), tipo_servico, "finalidade"):
        print(f"[AVISO] Não selecionou TIPO SERVIÇO='{tipo_servico}'")

    turno_el = wait.until(EC.presence_of_element_located((By.NAME, "turno")))
    if not _selecionar_option_relaxado(Select(turno_el), turno, "turno"):
        print(f"[AVISO] Não selecionou TURNO='{turno}'")

    botao = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "form button.btn.btn-success")))
    botao.click()

    _tratar_alerta_sucesso(driver)
    time.sleep(2)


# ---------------------------------------------------------------------------
# Função pública
# ---------------------------------------------------------------------------

def _matar_chrome():
    """Mata processos Chrome/Chromium órfãos antes de iniciar nova instância."""
    try:
        subprocess.run(["pkill", "-9", "-f", "chromium"], capture_output=True)
        subprocess.run(["pkill", "-9", "-f", "chromedriver"], capture_output=True)
        time.sleep(2)
    except Exception:
        pass


def _anexar_um_registro(nota_id: str, dados: Dict, idx: int, total: int,
                         usuario: str, senha: str) -> str | None:
    """
    Abre um Chrome, loga, anexa UM registro e fecha o Chrome.
    Retorna None em caso de sucesso, ou mensagem de erro.
    """
    _matar_chrome()
    driver = criar_driver()
    try:
        login_completo(driver, usuario, senha)

        if not _ir_para_anexos(driver, nota_id):
            return "Falha ao navegar para página de anexos"

        _preencher_registro(driver, dados, idx, total)
        return None
    except Exception as e:
        return str(e)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        time.sleep(3)  # pausa entre registros para liberar recursos


def anexar_lote(nota_id: str, xlsx_path: str, usuario: str, senha: str) -> Dict:
    """
    Anexa todos os registros de um XLSX a uma nota específica no SGI.
    Cada registro usa um Chrome próprio (cria → usa → destrói) para
    evitar crashes de memória em sessões longas.

    Retorna:
      {
        "nota_id": str,
        "total": int,
        "sucesso": int,
        "falhas": [{"indice": n, "matricula": str, "erro": str}]
      }
    """
    registros = _ler_registros_xlsx(xlsx_path)
    if not registros:
        return {
            "nota_id": nota_id,
            "total": 0,
            "sucesso": 0,
            "falhas": [],
            "erro": "Nenhum registro encontrado na planilha",
        }

    total = len(registros)
    sucesso = 0
    falhas = []

    for idx, dados in enumerate(registros, start=1):
        print(f"\n--- Registro {idx}/{total} ---")
        erro = _anexar_um_registro(nota_id, dados, idx, total, usuario, senha)
        if erro:
            print(f"[ERRO] Registro {idx}: {erro}")
            falhas.append({
                "indice": idx,
                "matricula": dados.get("matricula", ""),
                "erro": erro,
            })
        else:
            sucesso += 1

    return {
        "nota_id": nota_id,
        "total": total,
        "sucesso": sucesso,
        "falhas": falhas,
    }
